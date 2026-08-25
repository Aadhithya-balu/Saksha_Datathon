"""Tests for legacy data ingestion routes (issue #139 M1/M2)."""
import io

import pytest
from openpyxl import Workbook

from app.auth.dependencies import get_current_user
from app.core.security import hash_password
from app.models.crime_category import CrimeCategory
from app.models.location import Location
from app.models.role import Role
from app.models.user import User

IMPORTS = "/api/v2/data-import"


@pytest.fixture
def analyst_client(client, db_session):
    role = db_session.query(Role).filter_by(name="crime_analyst").first()
    if role is None:
        role = Role(name="crime_analyst", description="Crime Analyst")
        db_session.add(role)
        db_session.flush()
    user = User(
        username="import-analyst",
        email="import-analyst@example.com",
        full_name="Import Analyst",
        hashed_password=hash_password("Password123!"),
        role_id=role.id,
        is_active=True,
    )
    db_session.add(user)
    db_session.commit()
    client.app.dependency_overrides[get_current_user] = lambda: user
    yield client, user
    client.app.dependency_overrides.pop(get_current_user, None)


@pytest.fixture
def seed_reference(db_session):
    category = CrimeCategory(name="Theft & Burglaries", section_code="IPC 379", severity="medium")
    location = Location(district="Bengaluru Urban", station="KR Puram", latitude=13.0, longitude=77.7)
    db_session.add_all([category, location])
    db_session.commit()
    return {"category": category, "location": location}


def _csv_bytes(rows: list[str]) -> bytes:
    return ("\ufeff" + "\n".join(rows)).encode("utf-8")


VALID_VICTIMS_CSV = _csv_bytes([
    "full_name,gender,age,contact_number,address,statement",
    "Test Victim One,Male,30,9880000001,12 Main Rd Bengaluru,Snatched near bus stand",
    "test victim two,Female,,9880000002,Hassan,",
])


def test_list_entities(analyst_client):
    c, _ = analyst_client
    r = c.get(f"{IMPORTS}/entities")
    assert r.status_code == 200, r.text
    body = r.json()
    entity_types = {e["entity_type"] for e in body["entities"]}
    assert {"victims", "criminals", "crime_cases"} <= entity_types
    profiles = {p["profile"] for p in body["profiles"]}
    assert {"standard", "cctns"} <= profiles


@pytest.mark.parametrize("export_format", ["csv", "xlsx"])
def test_template_download(analyst_client, export_format):
    c, _ = analyst_client
    r = c.get(f"{IMPORTS}/template/victims?export_format={export_format}")
    assert r.status_code == 200, r.text
    assert len(r.content) > 50
    if export_format == "xlsx":
        assert r.headers["content-type"].startswith("application/vnd.openxmlformats")


def test_preview_validates_rows(analyst_client):
    """Preview must catch missing required fields and bad types without writing anything."""
    c, _ = analyst_client
    csv_file = _csv_bytes([
        "full_name,age",
        "Broken Row,not_a_number",
    ])
    r = c.post(
        f"{IMPORTS}/preview",
        files={"file": ("victims.csv", csv_file, "text/csv")},
        data={"entity_type": "victims", "profile": "standard"},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["total_rows"] == 1
    assert body["estimated_invalid_rows"] >= 1
    report_items = body["validation_report"]
    all_errors = [err for item in report_items for err in item["errors"]]
    assert any("age" in err for err in all_errors)
    assert any("integer" in err for err in all_errors)


def test_preview_auto_maps_headers(analyst_client):
    c, _ = analyst_client
    csv_file = _csv_bytes([
        "Full Name,Age",
        "Some Person,44",
    ])
    r = c.post(
        f"{IMPORTS}/preview",
        files={"file": ("victims.csv", csv_file, "text/csv")},
        data={"entity_type": "victims", "profile": "standard"},
    )
    body = r.json()
    assert body["column_mapping"]["Full Name"] == "full_name"
    assert body["column_mapping"]["Age"] == "age"
    assert body["unmapped_headers"] == []


def test_commit_imports_victims(analyst_client):
    c, _ = analyst_client
    r = c.post(
        f"{IMPORTS}/commit",
        files={"file": ("victims.csv", VALID_VICTIMS_CSV, "text/csv")},
        data={"entity_type": "victims", "profile": "standard"},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["total_rows"] == 2
    assert body["imported_rows"] == 2
    assert body["failed_rows"] == 0
    assert body["status"] == "completed"

    jobs = c.get(f"{IMPORTS}/jobs").json()
    assert jobs["total"] >= 1


def test_commit_crime_cases_with_relations(analyst_client, seed_reference):
    c, _ = analyst_client
    csv_file = _csv_bytes([
        "case_number,category_name,district,station,occurred_at,status,priority",
        "CR-TEST-9001,Theft & Burglaries,Bengaluru Urban,KR Puram,2026-07-14 22:30,open,high",
        "CR-TEST-9002,theft and burglaries,bangalore,KR Puram,2026/08/01,open,medium",
        "CR-TEST-9003,Unknown Category,Nowhere Land,2026-99-99,open",
    ])
    r = c.post(
        f"{IMPORTS}/commit",
        files={"file": ("cases.csv", csv_file, "text/csv")},
        data={"entity_type": "crime_cases", "profile": "standard"},
    )
    body = r.json()
    assert body["imported_rows"] == 2
    assert body["failed_rows"] == 1
    # Issue 5: partial success is reported via completed_with_warnings.
    assert body["status"] == "completed_with_warnings"
    errors = [err for item in body["validation_report"] for err in item["errors"]]
    assert any("category_name" in err for err in errors)


def test_cctns_profile_maps_headers(analyst_client, seed_reference):
    """M2: CCTNS extract headers map onto Saksha columns automatically."""
    c, _ = analyst_client
    csv_file = _csv_bytes([
        "REGISTRATION_NO,CRIME_HEAD,DISTRICT_NAME,POLICE_STATION,DATE_OF_REGISTRATION,FIR_STATUS",
        "CR-CCTNS-0001,Theft & Burglaries,Bengaluru Urban,KR Puram,2026-06-10,Open",
    ])
    preview = c.post(
        f"{IMPORTS}/preview",
        files={"file": ("cctns_extract.csv", csv_file, "text/csv")},
        data={"entity_type": "crime_cases", "profile": "cctns"},
    ).json()
    mapping = preview["column_mapping"]
    assert mapping["REGISTRATION_NO"] == "case_number"
    assert mapping["DISTRICT_NAME"] == "district"
    assert mapping["DATE_OF_REGISTRATION"] == "occurred_at"

    commit = c.post(
        f"{IMPORTS}/commit",
        files={"file": ("cctns_extract.csv", csv_file, "text/csv")},
        data={"entity_type": "crime_cases", "profile": "cctns"},
    ).json()
    assert commit["imported_rows"] == 1


def test_xlsx_upload_roundtrip(analyst_client):
    """M1: native .xlsx ingestion works end to end."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["full_name", "gender", "age"])
    sheet.append(["Excel Victim A", "Male", "41"])
    sheet.append(["Excel Victim B", "Female", "28"])
    buffer = io.BytesIO()
    workbook.save(buffer)

    c, _ = analyst_client
    r = c.post(
        f"{IMPORTS}/commit",
        files={"file": ("victims.xlsx", buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"entity_type": "victims", "profile": "standard"},
    )
    body = r.json()
    assert body["imported_rows"] == 2
    job_detail = c.get(f"{IMPORTS}/jobs/{body['job_id']}").json()
    assert job_detail["source_format"] == "xlsx"


def test_reports_xlsx_export(analyst_client):
    """M1: reports gain the previously-unimplemented xlsx export."""
    c, _ = analyst_client
    r = c.get("/api/v2/reports/criminals/export/xlsx")
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    from openpyxl import load_workbook
    loaded = load_workbook(io.BytesIO(r.content))
    assert loaded.active.title == "Report"
