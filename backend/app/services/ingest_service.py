"""Data ingestion service — bulk CSV/XLSX import with the full quality pipeline.

Supports:

- ``standard`` profile  : Saksha-native column templates (downloadable).
- ``cctns`` profile     : maps Crime and Criminal Tracking Network & Systems
                          (CCTNS) extract column headers onto Saksha entities,
                          so state CCTNS dumps can be ingested without manual
                          re-keying. See CCTNS_ICJS_INTEROP.md at the repo root.

Pipeline (issue 5, P1): file validation -> parse -> header mapping
-> normalization -> row validation -> staging -> duplicate detection
-> reconciliation vs trusted records -> quality grading -> admin promotion
into production tables with full source provenance. A parsed CSV row is never
trusted operational data until it is promoted. See INGESTION_PIPELINE.md.
"""
from __future__ import annotations

import csv
import io
import json
from datetime import date, datetime
from typing import Any

from sqlalchemy.orm import Session

# ---------------------------------------------------------------------------
# Column templates per entity (standard Saksha profile).
# Each field spec: column name -> {required, type, choices}
# ---------------------------------------------------------------------------

def _spec(required: bool = False, ftype: str = "string", choices: list[str] | None = None) -> dict[str, Any]:
    return {"required": required, "type": ftype, "choices": choices}


ENTITY_SPECS: dict[str, dict[str, dict[str, Any]]] = {
    "victims": {
        "full_name": _spec(required=True),
        "gender": _spec(choices=["Male", "Female", "Other"]),
        "age": _spec(ftype="integer"),
        "contact_number": _spec(),
        "address": _spec(),
        "statement": _spec(),
    },
    "criminals": {
        "full_name": _spec(required=True),
        "aliases": _spec(),
        "date_of_birth": _spec(ftype="date"),
        "gender": _spec(choices=["Male", "Female", "Other"]),
        "address": _spec(),
        "identifying_marks": _spec(),
        "mo_summary": _spec(),
        "status": _spec(choices=["at_large", "arrested", "convicted", "deceased"]),
    },
    "crime_cases": {
        "case_number": _spec(required=True),
        "category_name": _spec(required=True),
        "district": _spec(required=True),
        "station": _spec(),
        "occurred_at": _spec(required=True, ftype="datetime"),
        "description": _spec(),
        "mo_tags": _spec(),
        "status": _spec(choices=["open", "under_investigation", "closed", "convicted"]),
        "priority": _spec(choices=["low", "medium", "high", "critical"]),
        "progress": _spec(ftype="integer"),
    },
}

# ---------------------------------------------------------------------------
# M2: CCTNS extract header mapping. Keys are common CCTNS/ICJS column names
# (upper-snake, as produced by CCTNS "Search Arrested Person" / "Crime Details"
# exports); values map to the standard Saksha columns above. Unmapped CCTNS
# headers are reported as ignored so nothing is silently dropped.
# ---------------------------------------------------------------------------

CCTNS_COLUMN_MAPS: dict[str, dict[str, str]] = {
    "victims": {
        "NAME_OF_VICTIM": "full_name",
        "VICTIM_NAME": "full_name",
        "GENDER": "gender",
        "AGE": "age",
        "MOBILE_NUMBER": "contact_number",
        "CONTACT_NO": "contact_number",
        "ADDRESS": "address",
        "STATEMENT_DETAILS": "statement",
        "STATEMENT": "statement",
    },
    "criminals": {
        "ARRESTED_PERSON_NAME": "full_name",
        "PERSON_NAME": "full_name",
        "ACCUSED_NAME": "full_name",
        "ALIAS_NAME": "aliases",
        "ALIASES": "aliases",
        "DATE_OF_BIRTH": "date_of_birth",
        "DOB": "date_of_birth",
        "GENDER": "gender",
        "ADDRESS": "address",
        "IDENTIFICATION_MARKS": "identifying_marks",
        "MODUS_OPERANDI": "mo_summary",
        "MO_DETAILS": "mo_summary",
        "ARREST_STATUS": "status",
        "PERSON_STATUS": "status",
    },
    "crime_cases": {
        "REGISTRATION_NO": "case_number",
        "FIR_NO": "case_number",
        "CRIME_HEAD": "category_name",
        "IPC_SECTIONS": "category_name",
        "BNS_SECTIONS": "category_name",
        "DISTRICT_NAME": "district",
        "DISTRICT_CD": "district",
        "POLICE_STATION": "station",
        "PS_NAME": "station",
        "DATE_OF_REGISTRATION": "occurred_at",
        "INCIDENT_DATE": "occurred_at",
        "COMPLAINT_DETAILS": "description",
        "GENERAL_REMARKS": "description",
        "FIR_STATUS": "status",
        "CASE_STATUS": "status",
    },
}

VALID_IMPORT_ENTITIES = sorted(ENTITY_SPECS.keys())
VALID_PROFILES = ("standard", "cctns")

MAX_ROWS = 5000


class IngestError(Exception):
    """Raised when a file cannot be parsed at all."""


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

def parse_tabular_file(content: bytes, filename: str) -> tuple[list[str], list[dict[str, Any]]]:
    """Parse an uploaded CSV or XLSX file into (headers, row dicts).

    Rows are returned keyed by their original (stripped) header names; empty
    fully-blank rows are skipped.
    """
    lower = filename.lower()
    if lower.endswith((".xlsx", ".xls")):
        return _parse_xlsx(content)
    return _parse_csv(content)


def _parse_csv(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows_raw = [row for row in reader if any((cell or "").strip() for cell in row)]
    if not rows_raw:
        raise IngestError("File contains no data rows")
    headers = [h.strip() for h in rows_raw[0]]
    parsed: list[dict[str, Any]] = []
    for raw in rows_raw[1:]:
        row: dict[str, Any] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            row[header] = (raw[idx].strip() if idx < len(raw) else "")
        parsed.append(row)
    return headers, parsed


def _parse_xlsx(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - openpyxl is in requirements
        raise IngestError("Excel support is not installed (openpyxl missing)") from exc
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise IngestError(f"Could not read Excel file: {exc}") from exc
    sheet = workbook.active
    if sheet is None:
        raise IngestError("Excel workbook has no active sheet")
    matrix = [list(row) for row in sheet.iter_rows(values_only=True)]
    workbook.close()
    # Trim to contiguous non-empty region and normalize cells to strings.
    matrix = [row for row in matrix if any(cell is not None and str(cell).strip() for cell in row)]
    if not matrix:
        raise IngestError("Excel sheet contains no data rows")
    headers = [str(c).strip() if c is not None else "" for c in matrix[0]]
    parsed: list[dict[str, Any]] = []
    for raw in matrix[1:]:
        row = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            cell = raw[idx] if idx < len(raw) else None
            row[header] = "" if cell is None else str(cell).strip()
        parsed.append(row)
    return headers, parsed


# ---------------------------------------------------------------------------
# Header mapping (profile-aware, with fuzzy fallback)
# ---------------------------------------------------------------------------

def _normalize_header(header: str) -> str:
    return "".join(ch for ch in header.lower().replace(" ", "_") if ch.isalnum() or ch == "_").strip("_")


def build_column_mapping(headers: list[str], entity_type: str, profile: str) -> tuple[dict[str, str], list[str]]:
    """Map uploaded headers -> standard entity columns.

    Returns ``(mapping, unmapped_headers)`` where *mapping* keys are uploaded
    headers and values are standard column names. Unknown headers are listed
    in *unmapped_headers* so operators see what was skipped.
    """
    spec_columns = set(ENTITY_SPECS[entity_type].keys())
    normalized_targets = {_normalize_header(col): col for col in spec_columns}
    profile_map = {}
    if profile == "cctns":
        profile_map = {k.upper(): v for k, v in CCTNS_COLUMN_MAPS.get(entity_type, {}).items()}

    mapping: dict[str, str] = {}
    unmapped: list[str] = []
    for header in headers:
        if not header:
            continue
        if header in profile_map and profile_map[header] in spec_columns:
            mapping[header] = profile_map[header]
            continue
        normalized = _normalize_header(header)
        upper = header.upper()
        if upper in profile_map and profile_map[upper] in spec_columns:
            mapping[header] = profile_map[upper]
            continue
        if normalized in normalized_targets:
            mapping[header] = normalized_targets[normalized]
        else:
            unmapped.append(header)
    return mapping, unmapped


# ---------------------------------------------------------------------------
# Validation + coercion
# ---------------------------------------------------------------------------

def _coerce_value(raw: Any, ftype: str) -> tuple[Any, tuple[str, str] | None]:
    """Coerce a raw cell to *ftype*. Returns ``(value, (code, message) | None)``."""
    value = ("" if raw is None else str(raw)).strip()
    if not value:
        return None, None
    if ftype == "integer":
        try:
            return int(float(value)), None
        except ValueError:
            return None, ("INVALID_TYPE", f"'{value}' is not a valid integer")
    if ftype == "date":
        parsed = _parse_date(value)
        if parsed is None:
            return None, ("INVALID_DATE", f"'{value}' is not a recognizable date (use YYYY-MM-DD)")
        return parsed, None
    if ftype == "datetime":
        parsed = _parse_datetime(value)
        if parsed is None:
            return None, ("INVALID_DATETIME", f"'{value}' is not a recognizable datetime (use YYYY-MM-DD HH:MM)")
        return parsed, None
    return value, None


def _parse_date(value: str) -> date | None:
    candidates = ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d.%m.%Y", "%Y/%m/%d")
    for fmt in candidates:
        try:
            return datetime.strptime(value.split(" ")[0], fmt).date()
        except ValueError:
            continue
    return None


def _parse_datetime(value: str) -> datetime | None:
    candidates = (
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M",
        "%d-%m-%Y %H:%M", "%d/%m/%Y %H:%M", "%Y-%m-%d",
    )
    for fmt in candidates:
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    parsed_date = _parse_date(value)
    if parsed_date:
        return datetime(parsed_date.year, parsed_date.month, parsed_date.day)
    return None


def validate_row(
    db: Session,
    entity_type: str,
    standard_row: dict[str, Any],
    check_duplicates: bool = True,
    row_index: int | None = None,
) -> tuple[dict[str, Any], list[str], list[str]]:
    """Validate one mapped row. Returns (clean_values, errors, warnings).

    Errors/warnings use coded ``CODE(field): message`` strings so reports are
    machine-filterable while staying human-readable (issue 5 §9). The full
    pipeline passes ``check_duplicates=False`` and performs its own structured
    duplicate + reconciliation passes instead of this coarse inline check.
    """
    spec = ENTITY_SPECS[entity_type]
    clean: dict[str, Any] = {}
    errors: list[str] = []
    warnings: list[str] = []

    for column, rules in spec.items():
        raw = standard_row.get(column)
        coerced, err = _coerce_value(raw, rules["type"])
        if err:
            errors.append(f"{err[0]}({column}): {err[1]}")
            continue
        if coerced is None and rules["required"]:
            errors.append(f"REQUIRED_FIELD_MISSING({column}): required field is missing")
            continue
        if coerced is not None and rules.get("choices"):
            normalized = str(coerced).strip().lower()
            if normalized in rules["choices"]:
                coerced = normalized
            elif str(coerced) in rules["choices"]:
                pass
            else:
                errors.append(f"INVALID_CHOICE({column}): '{coerced}' must be one of {', '.join(rules['choices'])}")
                continue
        # Deterministic range sanity (documented normalization rules, issue 5 §7/§8).
        if coerced is not None and rules["type"] == "integer" and column == "age" and not (0 <= int(coerced) <= 120):
            warnings.append(f"OUT_OF_RANGE({column}): age '{coerced}' outside plausible range 0-120")
        clean[column] = coerced

    # Entity-specific relational validation. Always runs so reports surface
    # every problem per row, not just the first failure class.
    if entity_type == "crime_cases":
        category_name = clean.get("category_name")
        district = clean.get("district")
        station = clean.get("station")
        occurred_at = clean.get("occurred_at")
        if occurred_at is not None:
            from datetime import datetime as _dt
            if occurred_at > _dt.now():
                warnings.append(f"FUTURE_DATE(occurred_at): incident timestamp '{occurred_at.isoformat()}' is in the future")
        if category_name:
            from app.models.crime_category import CrimeCategory
            found = _match_category(db, CrimeCategory, category_name)
            if found:
                clean["category_id"] = found
            else:
                known = ", ".join(r[0] for r in db.query(CrimeCategory.name).limit(12).all())
                errors.append(f"UNKNOWN_CATEGORY(category_name): '{category_name}' not found (known: {known})")
        if district:
            from app.services.sociological_service import KARNATAKA_DISTRICTS
            matched = _match_district(district)
            if matched:
                clean["district"] = matched
            elif district not in KARNATAKA_DISTRICTS:
                warnings.append(f"DISTRICT_UNRECOGNIZED(district): '{district}' is not a recognised Karnataka reference district")
        if station and district:
            from app.models.location import Location
            # Station must actually match — never silently attach a row to an
            # arbitrary location in the district (issue 5 §8/§9).
            location = (
                db.query(Location)
                .filter(Location.district == clean.get("district", district))
                .filter(Location.station.ilike(station.strip()))
                .first()
            )
            if location:
                clean["location_id"] = location.id
                clean.pop("station", None)
                clean.pop("district", None)
            else:
                errors.append(f"LOCATION_NOT_FOUND(station): no location found for '{station}' in district '{district}'")

    # Duplicate detection (preview path only — the full pipeline uses
    # structured dedup + reconciliation instead, issue 5 §10-§13).
    if check_duplicates and entity_type == "crime_cases" and clean.get("case_number"):
        from app.models.crime import CrimeCase
        exists = db.query(CrimeCase.id).filter(CrimeCase.case_number == clean["case_number"]).first()
        if exists:
            errors.append(f"DUPLICATE_CASE_NUMBER(case_number): '{clean['case_number']}' already exists (duplicate)")
    if check_duplicates and entity_type == "criminals" and clean.get("full_name"):
        from app.models.criminal import Criminal
        exists = db.query(Criminal.id).filter(Criminal.full_name.ilike(clean["full_name"])).first()
        if exists:
            warnings.append(f"POSSIBLE_DUPLICATE(full_name): a criminal named '{clean['full_name']}' already exists; review for duplicates")

    return clean, errors, warnings


def _match_category(db: Session, CrimeCategory, raw: str) -> Any | None:
    """Resolve a category name tolerantly (case, '&' vs 'and', spacing)."""
    normalized = " ".join(raw.lower().replace("&", "and").split())
    for category in db.query(CrimeCategory).all():
        if category.name.strip().lower() == raw.strip().lower():
            return category.id
        candidate_norm = " ".join(category.name.lower().replace("&", "and").split())
        if candidate_norm == normalized:
            return category.id
    return None


_DISTRICT_ALIASES = {
    "bengaluru": "Bengaluru Urban",
    "bangalore": "Bengaluru Urban",
    "bengaluru urban": "Bengaluru Urban",
    "bangalore urban": "Bengaluru Urban",
    "mangalore": "Mangaluru",
    "dakshina kannada": "Mangaluru",
    "bellary": "Ballari",
    "gulbarga": "Kalaburagi",
    "kalaburagi": "Kalaburagi",
    "tumkur": "Tumkuru",
    "tumakuru": "Tumkuru",
    "hubli-dharwad": "Dharwad",
    "dharwad": "Dharwad",
}


def _match_district(raw: str) -> str | None:
    """Resolve a district string against known Karnataka districts + aliases."""
    from app.services.sociological_service import KARNATAKA_DISTRICTS
    needle = raw.strip().lower()
    if needle in _DISTRICT_ALIASES:
        candidate = _DISTRICT_ALIASES[needle]
        return candidate if candidate in KARNATAKA_DISTRICTS else None
    for name in KARNATAKA_DISTRICTS:
        if name.lower() == needle:
            return name
    return None


# ---------------------------------------------------------------------------
# High-level operations used by routes
# ---------------------------------------------------------------------------

def analyze_upload(
    db: Session,
    content: bytes,
    filename: str,
    entity_type: str,
    profile: str = "standard",
    limit_report_rows: int = 50,
) -> dict[str, Any]:
    """Parse + map + validate an upload WITHOUT writing anything.

    Returns the full preview payload: detected headers, proposed mapping,
    unmapped headers, sample mapped rows, and the validation report.
    """
    if entity_type not in ENTITY_SPECS:
        raise IngestError(f"Unsupported entity type '{entity_type}'")
    if profile not in VALID_PROFILES:
        raise IngestError(f"Unsupported mapping profile '{profile}'")

    headers, parsed_rows = parse_tabular_file(content, filename)
    if len(parsed_rows) > MAX_ROWS:
        raise IngestError(f"File has {len(parsed_rows)} rows; maximum supported is {MAX_ROWS}")

    mapping, unmapped = build_column_mapping(headers, entity_type, profile)
    missing_required = [
        col for col, rules in ENTITY_SPECS[entity_type].items() if rules["required"] and col not in mapping.values()
    ]

    report: list[dict[str, Any]] = []
    samples: list[dict[str, Any]] = []
    valid_count = 0
    for index, raw_row in enumerate(parsed_rows, start=2):  # +1 header line, +1 for 1-based
        standard_row = {mapping[h]: v for h, v in raw_row.items() if h in mapping}
        _, errors, warnings = validate_row(db, entity_type, standard_row)
        if not errors:
            valid_count += 1
        if len(report) < limit_report_rows and (errors or warnings):
            report.append({"row_number": index, "errors": errors, "warnings": warnings})
        if len(samples) < 5:
            samples.append(standard_row)

    error_count = sum(1 for item in report for e in item["errors"])  # only counted within sampled slice
    return {
        "entity_type": entity_type,
        "profile": profile,
        "filename": filename,
        "detected_headers": headers,
        "column_mapping": mapping,
        "unmapped_headers": unmapped,
        "missing_required_columns": missing_required,
        "total_rows": len(parsed_rows),
        "sample_mapped_rows": samples,
        "validation_report": report,
        "truncated_report": max(0, len(parsed_rows) - len(report)) > 0,
        "estimated_valid_rows": valid_count,
        "estimated_invalid_rows": len(parsed_rows) - valid_count,
        "_diagnostic_error_sample_count": error_count,
    }


def commit_import(
    db: Session,
    content: bytes,
    filename: str,
    entity_type: str,
    profile: str,
    created_by_id,
) -> dict[str, Any]:
    """Validate then persist all valid rows. Returns the final import report."""
    preview = analyze_upload(db, content, filename, entity_type, profile, limit_report_rows=10**9)

    if entity_type == "victims":
        from app.models.victim import Victim as model
    elif entity_type == "criminals":
        from app.models.criminal import Criminal as model
    else:
        from app.models.crime import CrimeCase as model

    headers, parsed_rows = parse_tabular_file(content, filename)
    mapping, _unmapped = build_column_mapping(headers, entity_type, profile)

    # Virtual columns used for lookup/validation; never passed to the ORM.
    _VIRTUAL_COLUMNS = {"category_name", "district", "station"}

    imported = 0
    full_report: list[dict[str, Any]] = []
    for index, raw_row in enumerate(parsed_rows, start=2):
        standard_row = {mapping[h]: v for h, v in raw_row.items() if h in mapping}
        clean, errors, warnings = validate_row(db, entity_type, standard_row)
        if errors:
            full_report.append({"row_number": index, "errors": errors, "warnings": warnings})
            continue
        if entity_type == "crime_cases":
            clean.setdefault("status", "open")
            clean.setdefault("priority", "medium")
            clean.setdefault("progress", 10)
            clean = {k: v for k, v in clean.items() if k not in _VIRTUAL_COLUMNS}
        obj = model(**{k: v for k, v in clean.items()})
        db.add(obj)
        imported += 1
        if warnings:
            full_report.append({"row_number": index, "errors": [], "warnings": warnings})

    failed = len(parsed_rows) - imported
    status = "completed" if failed == 0 else ("partial" if imported > 0 else "failed")

    job_record = {
        "entity_type": entity_type,
        "source_format": "xlsx" if filename.lower().endswith((".xlsx", ".xls")) else "csv",
        "mapping_profile": profile,
        "filename": filename,
        "total_rows": len(parsed_rows),
        "imported_rows": imported,
        "failed_rows": failed,
        "status": status,
        "validation_report": json.dumps(full_report[:200]),
        "created_by_id": created_by_id,
    }
    return {"job": job_record, "report": full_report, "preview": preview}


def build_template(entity_type: str, export_format: str = "csv") -> tuple[bytes, str, str]:
    """Build a downloadable import template. Returns (bytes, media_type, extension)."""
    columns = list(ENTITY_SPECS[entity_type].keys())
    example_rows = _TEMPLATE_EXAMPLES.get(entity_type, [])
    if export_format == "xlsx":
        from openpyxl import Workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(columns)
        for row in example_rows:
            sheet.append([row.get(c, "") for c in columns])
        buffer = io.BytesIO()
        workbook.save(buffer)
        return (
            buffer.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "xlsx",
        )
    buffer = io.StringIO()
    buffer.write("\ufeff")
    writer = csv.DictWriter(buffer, fieldnames=columns)
    writer.writeheader()
    for row in example_rows:
        writer.writerow({c: row.get(c, "") for c in columns})
    return buffer.getvalue().encode("utf-8"), "text/csv; charset=utf-8", "csv"


_TEMPLATE_EXAMPLES: dict[str, list[dict[str, str]]] = {
    "victims": [
        {
            "full_name": "Suresh Gowda",
            "gender": "Male",
            "age": "34",
            "contact_number": "9880012345",
            "address": "12, 4th Cross, KR Puram, Bengaluru Urban",
            "statement": "Two men snatched his bag near the bus stand at 21:30.",
        },
        {
            "full_name": "Lakshmi Rao",
            "gender": "Female",
            "age": "67",
            "contact_number": "",
            "address": "Hassan town",
            "statement": "",
        },
    ],
    "criminals": [
        {
            "full_name": "Ravi alias Kulla",
            "aliases": "Kulla, Ravi M",
            "date_of_birth": "1992-05-14",
            "gender": "Male",
            "address": "Ballari",
            "identifying_marks": "Scar over left eyebrow",
            "mo_summary": "Targets unlocked two-wheelers at night; uses stolen vehicles KA-35-AB-1234.",
            "status": "at_large",
        },
    ],
    "crime_cases": [
        {
            "case_number": "CR-2026-9001",
            "category_name": "Theft & Burglaries",
            "district": "Bengaluru Urban",
            "station": "KR Puram",
            "occurred_at": "2026-07-14 22:30",
            "description": "Housebreak at night; gold ornaments stolen.",
            "mo_tags": "night_entry,locked_house",
            "status": "open",
            "priority": "high",
            "progress": "25",
        },
    ],
}


# ===========================================================================
# Issue 5 (P1): full ingestion pipeline.
#
#   upload -> import job -> staging -> mapping -> normalization -> validation
#          -> deduplication -> reconciliation -> quality grading
#          -> admin promotion -> trusted Saksha records (with provenance)
#
# A successfully parsed CSV row is NEVER treated as trusted operational
# intelligence on its own: it must land in import_staging_records as a
# NEW_RECORD and be explicitly promoted by an administrator before it reaches
# production tables. See INGESTION_PIPELINE.md at the repo root.
# ===========================================================================

import uuid as _uuid  # noqa: E402
from difflib import SequenceMatcher  # noqa: E402

from app.models.import_job import ImportJob, ImportStagedRecord  # noqa: E402

MAX_FILE_BYTES = 10 * 1024 * 1024  # 10 MB hard cap before any parsing happens

# Documented, configurable quality-grade thresholds (issue 5 §16). A row counts
# towards `problem_ratio` when it is invalid or conflicts with trusted data;
# warnings/duplicates degrade confidence but do not dominate the grade.
GRADE_THRESHOLDS = {
    "A": 0.02,   # <= 2% problem rows and no conflicts
    "B": 0.10,   # <= 10% problem rows
    "C": 0.25,   # <= 25% problem rows
}
# Above C threshold, or zero valid rows -> D; zero valid + all rejected or
# >50% problems -> REJECTED (dataset is not safe to promote at all).

_JOB_STATUSES = (
    "uploaded", "processing", "validated", "reconciling",
    "completed", "completed_with_warnings", "failed", "cancelled",
)

_VIRTUAL_COLUMNS = {"category_name", "district", "station"}


class ImportSecurityError(IngestError):
    """Raised when an upload fails file-level security/validation checks."""


# ---------------------------------------------------------------------------
# File-level validation (issue 5 §24) — never trust names/extensions alone.
# ---------------------------------------------------------------------------

def validate_file(content: bytes, filename: str) -> str:
    """Validate size/type/content of an upload. Returns 'csv' or 'xlsx'.

    Rejects oversized files, unknown extensions, XLSX files whose bytes are not
    a real ZIP container (OOXML magic ``PK\\x03\\x04``), and CSVs that do not
    decode as UTF-8 text. Cell contents remain untrusted downstream.
    """
    lower = (filename or "").lower()
    if len(content) == 0:
        raise ImportSecurityError("Uploaded file is empty")
    if len(content) > MAX_FILE_BYTES:
        raise ImportSecurityError(f"File exceeds the {MAX_FILE_BYTES // (1024 * 1024)} MB upload limit")
    if lower.endswith((".xlsx", ".xls")):
        if not content.startswith(b"PK\x03\x04"):
            raise ImportSecurityError("File has an Excel extension but is not a valid XLSX workbook")
        return "xlsx"
    if not lower.endswith(".csv"):
        raise ImportSecurityError("Unsupported file type: only .csv and .xlsx uploads are accepted")
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ImportSecurityError("CSV file must be UTF-8 encoded") from exc
    # Binary/executable masquerading as CSV: NUL bytes or HTML script payloads.
    if "\x00" in text[:4096] or "<script" in text[:4096].lower():
        raise ImportSecurityError("File does not look like a valid CSV document")
    return "xlsx" if lower.endswith((".xlsx", ".xls")) else "csv"


# ---------------------------------------------------------------------------
# Deterministic normalization (issue 5 §7). Pure functions — no DB access, no
# locale dependence, documented in INGESTION_PIPELINE.md.
# ---------------------------------------------------------------------------

def normalize_text(value: Any) -> str:
    """Trim and collapse internal whitespace; preserve case of sensitive values."""
    return " ".join(("" if value is None else str(value)).split())


def normalize_identifier(value: Any) -> str:
    """Uppercase + collapse whitespace for stable identifiers (FIR/case numbers)."""
    return normalize_text(value).upper()


def normalize_phone(value: Any) -> str:
    """Strip separators from phone numbers, keeping leading '+'."""
    raw = normalize_text(value)
    if not raw:
        return ""
    keep_plus = raw.startswith("+")
    digits = "".join(ch for ch in raw if ch.isdigit())
    return ("+" if keep_plus else "") + digits


def normalize_person_key(value: Any) -> str:
    """Casefolded name key used ONLY for duplicate matching — never stored."""
    return normalize_text(value).casefold()


def normalize_choice_value(value: Any) -> str:
    return normalize_text(value).strip().lower()


_ROW_NORMALIZERS = {
    "case_number": normalize_identifier,
    "contact_number": normalize_phone,
}


def normalize_row(entity_type: str, standard_row: dict[str, Any]) -> dict[str, Any]:
    """Apply the documented per-column normalizers to one mapped row."""
    normalized = {}
    for column, value in standard_row.items():
        normalizer = _ROW_NORMALIZERS.get(column, normalize_text)
        normalized[column] = normalizer(value)
    return normalized


def _jsonable(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, _uuid.UUID):
        return str(value)
    return value


# ---------------------------------------------------------------------------
# Duplicate detection (§10/§11) + reconciliation (§13/§14).
# ---------------------------------------------------------------------------

def _person_duplicate_key(clean: dict[str, Any]) -> str | None:
    name = clean.get("full_name")
    return normalize_person_key(name) if name else None


def _is_strong_person_match(a: dict[str, Any], b: dict[str, Any]) -> bool:
    """Exact duplicate: same normalized name AND same DOB or contact number."""
    dob_a, dob_b = a.get("date_of_birth"), b.get("date_of_birth")
    contact_a, contact_b = a.get("contact_number"), b.get("contact_number")
    dob_equal = dob_a is not None and dob_b is not None and dob_a == dob_b
    contact_equal = bool(contact_a) and bool(contact_b) and contact_a == contact_b
    return dob_equal or contact_equal


def _fuzzy_name_similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a, b).ratio()


def _dedupe_and_reconcile(
    db: Session,
    entity_type: str,
    staged: list[dict[str, Any]],
) -> None:
    """Annotate staged rows in place with duplicate/reconciliation outcomes.

    Each entry of *staged* is ``{"clean": ..., "record": ImportStagedRecord}``.
    Outcomes (conservative by design — existing trusted records are never
    overwritten here):

    - crime_cases: exact key = case_number. DB match compares fields; identical
      -> MATCHED (skip), differing -> CONFLICT flagged for review with both
      values preserved; unseen -> NEW_RECORD (promotable).
    - criminals/victims: no stable external identifier exists in the schema, so
      name+DOB/contact agreement marks EXISTING_MATCH (skip), bare identical
      normalized names flag POTENTIAL_DUPLICATE (held for admin review).
    """
    from app.models.crime import CrimeCase
    from app.models.criminal import Criminal
    from app.models.victim import Victim

    if entity_type == "crime_cases":
        keys = [normalize_identifier(s["clean"].get("case_number") or "") for s in staged]
        candidate_keys = [k for k in keys if k]
        existing = {}
        if candidate_keys:
            for row in (
                db.query(CrimeCase)
                .filter(CrimeCase.case_number.in_(candidate_keys))
                .all()
            ):
                existing[row.case_number.upper()] = row
        seen_keys: dict[str, int] = {}
        for idx, item in enumerate(staged):
            record, clean = item["record"], item["clean"]
            key = keys[idx]
            if not key:
                _set_reconciliation(record, "new_record", "validated")
                continue
            if key in seen_keys:
                _set_reconciliation(record, "duplicate", "rejected")
                record.duplicate_status = "exact_duplicate"
                record.duplicate_of = _dumps([{"kind": "batch_row", "row_number": staged[seen_keys[key]]["record"].row_number, "key": key}])
                continue
            seen_keys[key] = idx
            match = existing.get(key)
            if match is None:
                _set_reconciliation(record, "new_record", "validated_with_warnings" if record.validation_status == "warning" else "validated")
                continue
            diffs = _compare_case(match, clean)
            if diffs:
                record.duplicate_status = "existing_match"
                _set_reconciliation(
                    record,
                    "conflict",
                    "review_required",
                    {"matched_case_id": str(match.id), "case_number": key, "field_conflicts": diffs},
                )
            else:
                record.duplicate_status = "existing_match"
                _set_reconciliation(
                    record,
                    "duplicate",
                    "validated",
                    {"matched_case_id": str(match.id), "case_number": key},
                    note="identical to trusted Saksha case — skipped",
                )
        return

    # Persons (criminals/victims): batch-internal + DB fuzzy detection.
    model = Criminal if entity_type == "criminals" else Victim
    names = [_person_duplicate_key(s["clean"]) for s in staged]
    unique_names = sorted({n for n in names if n})
    db_by_name = {}
    if unique_names:
        # Indexed-ish lookup in chunks; avoids one query per row (§30).
        from sqlalchemy import or_
        for start in range(0, len(unique_names), 200):
            chunk = unique_names[start:start + 200]
            conditions = [model.full_name.ilike(n) for n in chunk]
            for row in db.query(model).filter(or_(*conditions)).all():
                db_by_name[normalize_person_key(row.full_name)] = row
    seen: dict[str, dict[str, Any]] = {}
    for idx, item in enumerate(staged):
        record, clean = item["record"], item["clean"]
        name_key = names[idx]
        if not name_key:
            _set_reconciliation(record, "new_record", "validated_with_warnings" if record.validation_status == "warning" else "validated")
            continue
        batch_hit = seen.get(name_key)
        db_hit = db_by_name.get(name_key)
        strong_ref = None
        for candidate in (batch_hit, db_hit):
            if candidate is not None and _is_strong_person_match(clean, candidate["attrs"] if isinstance(candidate, dict) else _person_attrs(candidate)):
                strong_ref = candidate
                break
        if strong_ref is not None:
            ref_row_number = strong_ref["record"].row_number if isinstance(strong_ref, dict) else None
            ref_id = None if isinstance(strong_ref, dict) else str(strong_ref.id)
            record.duplicate_status = "exact_duplicate"
            _set_reconciliation(
                record, "duplicate", "rejected",
                {
                    "kind": "batch_row" if isinstance(strong_ref, dict) else "saksha_record",
                    "row_number": ref_row_number,
                    "id": ref_id,
                    "key": name_key,
                },
            )
            continue
        if batch_hit is not None or db_hit is not None:
            record.duplicate_status = "potential_duplicate"
            refs = []
            if batch_hit is not None:
                refs.append({"kind": "batch_row", "row_number": batch_hit["record"].row_number, "key": name_key})
            if db_hit is not None:
                refs.append({"kind": "saksha_record", "id": str(db_hit.id), "key": name_key})
            record.duplicate_of = _dumps(refs)
            _set_reconciliation(
                record, "review_required", "review_required",
                {"reason": "POTENTIAL_DUPLICATE", "refs": refs,
                 "note": "same normalized name as an existing person; held for admin review"},
            )
            continue
        seen[name_key] = {"record": record, "attrs": clean}
        _set_reconciliation(record, "new_record", "validated_with_warnings" if record.validation_status == "warning" else "validated")


def _person_attrs(row) -> dict[str, Any]:
    return {"full_name": row.full_name, "date_of_birth": getattr(row, "date_of_birth", None), "contact_number": getattr(row, "contact_number", None)}


def _compare_case(existing: Any, clean: dict[str, Any]) -> dict[str, dict[str, Any]]:
    """Field-by-field diff between a trusted case and an imported duplicate."""
    diffs: dict[str, dict[str, Any]] = {}

    def consider(field: str, existing_value: Any, imported_value: Any) -> None:
        if imported_value in (None, ""):
            return  # absent import values never count as conflicts
        existing_repr = _jsonable(existing_value)
        imported_repr = _jsonable(imported_value)
        if existing_repr != imported_repr:
            diffs[field] = {"existing": existing_repr, "imported": imported_repr}

    consider("status", existing.status, clean.get("status"))
    consider("priority", existing.priority, clean.get("priority"))
    consider("description", existing.description, clean.get("description"))
    if clean.get("occurred_at") is not None and existing.occurred_at is not None:
        imported_dt = clean["occurred_at"]
        existing_dt = existing.occurred_at.replace(tzinfo=None) if existing.occurred_at.tzinfo else existing.occurred_at
        if existing_dt != imported_dt:
            diffs["occurred_at"] = {"existing": _jsonable(existing.occurred_at), "imported": _jsonable(imported_dt)}
    return diffs


def _set_reconciliation(
    record: ImportStagedRecord,
    status: str,
    trust: str,
    details: dict[str, Any] | None = None,
    note: str | None = None,
) -> None:
    payload = dict(details or {})
    if note:
        payload["note"] = note
    record.reconciliation_status = status
    record.trust_level = trust
    record.reconciliation_details = _dumps(payload) if payload else None


def _dumps(payload: Any) -> str:
    return json.dumps(payload, default=_jsonable)


# ---------------------------------------------------------------------------
# Quality grading (§15/§16/§17) — computed strictly from job metrics.
# ---------------------------------------------------------------------------

def compute_quality_grade(metrics: dict[str, int]) -> str:
    """Grade an import from its actual validation metrics.

    problem_ratio = (invalid_rows + conflict_rows) / total_rows.
      A  <= 2%  and no conflicts          B  <= 10%
      C  <= 25%                           D  > 25% (or no valid rows)
      REJECTED  > 50% problems or zero valid rows on a non-empty dataset.
    """
    total = metrics.get("total_rows", 0)
    invalid = metrics.get("invalid_rows", 0)
    conflicts = metrics.get("conflict_rows", 0)
    valid = metrics.get("valid_rows", 0)
    if total == 0:
        return "REJECTED"
    if valid == 0:
        return "REJECTED"
    problem_ratio = (invalid + conflicts) / total
    if problem_ratio > 0.5:
        return "REJECTED"
    if problem_ratio > GRADE_THRESHOLDS["C"]:
        return "D"
    if problem_ratio > GRADE_THRESHOLDS["B"]:
        return "C"
    if problem_ratio > GRADE_THRESHOLDS["A"] or conflicts > 0:
        return "B"
    return "A"


# ---------------------------------------------------------------------------
# The pipeline itself (§3/§18/§19/§20).
# ---------------------------------------------------------------------------

def run_import_pipeline(
    db: Session,
    content: bytes,
    filename: str,
    entity_type: str,
    profile: str,
    created_by_id,
    source_system: str = "manual_upload",
) -> ImportJob:
    """Execute the complete ingestion pipeline for one upload.

    Parses nothing blindly: file validation runs first, every row is staged
    (never inserted into production tables), and the resulting ImportJob carries
    full metrics, quality grade and a row-level report. Promotion is a separate
    administrator-gated step (see :func:`promote_import`).
    """
    if entity_type not in ENTITY_SPECS:
        raise IngestError(f"Unsupported entity type '{entity_type}'")
    if profile not in VALID_PROFILES:
        raise IngestError(f"Unsupported mapping profile '{profile}'")

    source_format = validate_file(content, filename)

    job = ImportJob(
        entity_type=entity_type,
        source_format=source_format,
        mapping_profile=profile,
        source_system=source_system,
        filename=filename,
        status="processing",
        created_by_id=created_by_id,
        processing_started_at=datetime.now(),
    )
    db.add(job)
    db.flush()

    try:
        headers, parsed_rows = parse_tabular_file(content, filename)
        if len(parsed_rows) > MAX_ROWS:
            raise IngestError(f"File has {len(parsed_rows)} rows; maximum supported is {MAX_ROWS}")

        mapping, unmapped = build_column_mapping(headers, entity_type, profile)
        missing_required = [
            col for col, rules in ENTITY_SPECS[entity_type].items()
            if rules["required"] and col not in mapping.values()
        ]
        if missing_required and parsed_rows:
            raise IngestError(
                f"Import rejected: required Saksha columns have no source mapping: {', '.join(missing_required)}"
            )

        job.status = "reconciling"
        staged_items: list[dict[str, Any]] = []
        report: list[dict[str, Any]] = []

        for index, raw_row in enumerate(parsed_rows, start=1):
            spreadsheet_row = index + 1  # header offset for operator reference
            standard_row = normalize_row(entity_type, {mapping[h]: v for h, v in raw_row.items() if h in mapping})
            clean, errors, warnings = validate_row(db, entity_type, standard_row, check_duplicates=False)

            validation_status = "invalid" if errors else ("warning" if warnings else "valid")
            record = ImportStagedRecord(
                job_id=job.id,
                row_number=index,
                source_row_ref=str(spreadsheet_row),
                raw_data=_dumps(raw_row),
                # Validated/coerced values (incl. resolved category/location ids)
                # for promotion; verbatim source values stay in raw_data.
                mapped_data=_dumps(clean if not errors else standard_row),
                validation_status=validation_status,
                validation_errors=_dumps([_structured(*_parse_coded(e)) for e in errors]),
                validation_warnings=_dumps([_structured(*_parse_coded(w)) for w in warnings]),
                trust_level="rejected" if errors else "pending",
            )
            db.add(record)
            staged_items.append({"record": record, "clean": clean})

            if errors or warnings:
                report.append({
                    "row_number": spreadsheet_row,
                    "errors": errors,
                    "warnings": warnings,
                    # Structured mirrors of the coded strings above.
                    "error_details": [_structured(*_parse_coded(e)) for e in errors],
                    "warning_details": [_structured(*_parse_coded(w)) for w in warnings],
                    "validation_status": validation_status,
                })

        # Deduplication + reconciliation passes annotate the staged rows.
        job.status = "validated"
        db.flush()
        _dedupe_and_reconcile(db, entity_type, staged_items)
        job.status = "reconciling"
        # Invalid rows are never promotable regardless of their key matches.
        for item in staged_items:
            if item["record"].validation_status == "invalid":
                _set_reconciliation(item["record"], "rejected", "rejected")

        metrics = _compute_metrics(job, staged_items, unmapped_count=len(unmapped))
        job.quality_grade = compute_quality_grade(metrics)
        job.processing_completed_at = datetime.now()
        job.status = "completed" if (
            metrics["invalid_rows"] == 0
            and metrics["warning_rows"] == 0
            and metrics["conflict_rows"] == 0
            and metrics["potential_duplicate_rows"] == 0
        ) else "completed_with_warnings"
        job.validation_report = _dumps(report[:200])
    except Exception:
        job.status = "failed"
        job.processing_completed_at = datetime.now()
        job.quality_grade = "REJECTED"
        raise
    return job


def _parse_coded(message: str) -> tuple[str, str, str]:
    """Split ``CODE(field): rest`` into its parts; falls back gracefully."""
    head, _, rest = message.partition(": ")
    if "(" in head and head.endswith(")") and head != head.lower():
        code, field = head[:-1].split("(", 1)
        return code, field, rest
    return "VALIDATION_NOTE", "", message


def _structured(code: str, field: str, message: str) -> dict[str, str]:
    return {"code": code, "field": field, "message": message}


def _compute_metrics(
    job: ImportJob,
    staged_items: list[dict[str, Any]],
    unmapped_count: int = 0,
) -> dict[str, int]:
    """Derive the full quality metric set from the actual staged rows (§17)."""
    metrics = {
        "total_rows": len(staged_items),
        "valid_rows": 0,
        "invalid_rows": 0,
        "warning_rows": 0,
        "exact_duplicate_rows": 0,
        "potential_duplicate_rows": 0,
        "conflict_rows": 0,
        "new_record_rows": 0,
        "matched_record_rows": 0,
        "updated_record_rows": 0,
        "rejected_rows": 0,
        "review_rows": 0,
        "error_count": 0,
        "promoted_rows": 0,
        "unmapped_columns": unmapped_count,
    }
    error_count = 0
    for item in staged_items:
        record = item["record"]
        if record.validation_status == "valid":
            metrics["valid_rows"] += 1
        elif record.validation_status == "invalid":
            metrics["invalid_rows"] += 1
        elif record.validation_status == "warning":
            metrics["warning_rows"] += 1
            metrics["valid_rows"] += 1  # warnings are non-blocking
        if record.validation_errors:
            try:
                error_count += len(json.loads(record.validation_errors))
            except json.JSONDecodeError:
                pass
        dup = record.duplicate_status
        if dup == "exact_duplicate":
            metrics["exact_duplicate_rows"] += 1
        elif dup == "potential_duplicate":
            metrics["potential_duplicate_rows"] += 1
        recon = record.reconciliation_status
        if recon == "new_record":
            metrics["new_record_rows"] += 1
        elif recon == "matched":
            metrics["matched_record_rows"] += 1
        elif recon == "conflict":
            metrics["conflict_rows"] += 1
        elif recon == "duplicate":
            metrics["rejected_rows"] += 1
        elif recon == "review_required":
            metrics["review_rows"] += 1
        elif recon == "rejected":
            metrics["rejected_rows"] += 1
    metrics["error_count"] = error_count

    job.total_rows = metrics["total_rows"]
    job.valid_rows = metrics["valid_rows"]
    job.invalid_rows = metrics["invalid_rows"]
    job.warning_rows = metrics["warning_rows"]
    job.exact_duplicate_rows = metrics["exact_duplicate_rows"]
    job.potential_duplicate_rows = metrics["potential_duplicate_rows"]
    job.conflict_rows = metrics["conflict_rows"]
    job.new_record_rows = metrics["new_record_rows"]
    job.matched_record_rows = metrics["matched_record_rows"]
    job.updated_record_rows = metrics["updated_record_rows"]
    job.rejected_rows = metrics["rejected_rows"]
    job.review_rows = metrics["review_rows"]
    job.error_count = metrics["error_count"]
    job.failed_rows = metrics["invalid_rows"]
    job.imported_rows = metrics["new_record_rows"]
    return metrics


# ---------------------------------------------------------------------------
# Promotion (§18/§26) — the ONLY path from staging into trusted tables.
# ---------------------------------------------------------------------------

_PROMOTABLE_MODELS = {}


def _model_for(entity_type: str):
    if entity_type == "victims":
        from app.models.victim import Victim as model
    elif entity_type == "criminals":
        from app.models.criminal import Criminal as model
    elif entity_type == "crime_cases":
        from app.models.crime import CrimeCase as model
    else:
        raise IngestError(f"Unsupported entity type '{entity_type}'")
    return model


def promote_import(db: Session, job: ImportJob, promoted_by_id, include_review: bool = False) -> dict[str, Any]:
    """Promote eligible staged rows into production tables with provenance.

    Eligible = NEW_RECORD rows (plus POTENTIAL_DUPLICATE review rows when an
    admin explicitly passes ``include_review=True``). CONFLICT rows are never
    auto-promoted — the existing trusted record always wins until someone
    resolves the conflict deliberately. Runs inside a single transaction so a
    failure leaves no partial state behind (§20).
    """
    if job.status in ("failed", "cancelled"):
        raise IngestError("Cannot promote a failed or cancelled import job")
    if job.rolled_back_at is not None:
        raise IngestError("This import was rolled back and cannot be re-promoted")

    staged = (
        db.query(ImportStagedRecord)
        .filter(ImportStagedRecord.job_id == job.id, ImportStagedRecord.promoted.is_(False))
        .all()
    )
    eligible_statuses = {"new_record"}
    if include_review:
        eligible_statuses.add("review_required")

    promoted = 0
    skipped_conflicts = 0
    skipped_invalid = 0
    model = _model_for(job.entity_type)

    try:
        for record in staged:
            if record.reconciliation_status == "conflict":
                skipped_conflicts += 1
                continue
            if record.reconciliation_status not in eligible_statuses:
                continue
            if record.validation_status not in ("valid", "warning"):
                skipped_invalid += 1
                continue
            clean = json.loads(record.mapped_data) if record.mapped_data else {}
            fields = {k: v for k, v in clean.items() if k not in _VIRTUAL_COLUMNS}
            if job.entity_type == "crime_cases":
                fields.setdefault("status", "open")
                fields.setdefault("priority", "medium")
                fields["progress"] = int(fields.get("progress") or 10)
                # JSON staging round-trips dates/UUIDs as strings — restore types.
                if isinstance(fields.get("occurred_at"), str):
                    fields["occurred_at"] = datetime.fromisoformat(fields["occurred_at"])
                for key in ("category_id", "location_id"):
                    if isinstance(fields.get(key), str):
                        fields[key] = _uuid.UUID(fields[key])
            elif job.entity_type == "criminals" and isinstance(fields.get("date_of_birth"), str):
                fields["date_of_birth"] = datetime.fromisoformat(fields["date_of_birth"]).date()
            obj = model(
                **fields,
                dataset_provenance="migrated",
                source_import_job_id=job.id,
                source_file=job.filename,
                source_row_ref=record.source_row_ref,
            )
            db.add(obj)
            db.flush()
            record.promoted = True
            record.promoted_record_id = obj.id
            record.promoted_at = datetime.now()
            promoted += 1
        job.promoted_rows = (job.promoted_rows or 0) + promoted
        job.promoted_at = datetime.now()
        job.promoted_by_id = promoted_by_id
    except Exception:
        db.rollback()
        raise
    return {"promoted_rows": promoted, "skipped_conflicts": skipped_conflicts, "skipped_invalid": skipped_invalid}


def rollback_import(db: Session, job: ImportJob) -> int:
    """Undo a promotion: delete ONLY records this import created.

    Targets rows whose ``source_import_job_id`` equals this job, so unrelated
    operational data can never be touched (§20). Staged rows are marked
    unpromoted and the job is cancelled for traceability.
    """
    if job.rolled_back_at is not None:
        raise IngestError("Import has already been rolled back")
    model = _model_for(job.entity_type)
    records = db.query(model).filter(model.source_import_job_id == job.id).all()
    record_ids = {r.id for r in records}
    for obj in records:
        db.delete(obj)
    staged = db.query(ImportStagedRecord).filter(ImportStagedRecord.job_id == job.id).all()
    for record in staged:
        if record.promoted_record_id in record_ids:
            record.promoted = False
            record.promoted_record_id = None
            record.promoted_at = None
    job.promoted_rows = 0
    job.rolled_back_at = datetime.now()
    job.status = "cancelled"
    return len(records)


# ---------------------------------------------------------------------------
# Serialization helpers for the API layer.
# ---------------------------------------------------------------------------

def serialize_job(job: ImportJob, validation_report: list[dict[str, Any]] | None = None) -> dict[str, Any]:
    payload = {
        "id": str(job.id),
        "entity_type": job.entity_type,
        "source_format": job.source_format,
        "mapping_profile": job.mapping_profile,
        "profile": job.mapping_profile,  # legacy field name used by the admin panel
        "source_system": job.source_system,
        "filename": job.filename,
        "status": job.status,
        "quality_grade": job.quality_grade,
        "total_rows": job.total_rows,
        "valid_rows": job.valid_rows,
        "invalid_rows": job.invalid_rows,
        "warning_rows": job.warning_rows,
        "exact_duplicate_rows": job.exact_duplicate_rows,
        "potential_duplicate_rows": job.potential_duplicate_rows,
        "conflict_rows": job.conflict_rows,
        "new_record_rows": job.new_record_rows,
        "matched_record_rows": job.matched_record_rows,
        "updated_record_rows": job.updated_record_rows,
        "rejected_rows": job.rejected_rows,
        "review_rows": job.review_rows,
        "error_count": job.error_count,
        "promoted_rows": job.promoted_rows,
        "imported_rows": job.imported_rows,
        "failed_rows": job.failed_rows,
        "created_at": job.created_at.isoformat() if job.created_at else None,
        "processing_started_at": job.processing_started_at.isoformat() if job.processing_started_at else None,
        "processing_completed_at": job.processing_completed_at.isoformat() if job.processing_completed_at else None,
        "promoted_at": job.promoted_at.isoformat() if job.promoted_at else None,
        "rolled_back_at": job.rolled_back_at.isoformat() if job.rolled_back_at else None,
        "created_by": job.created_by.full_name if job.created_by else None,
    }
    if validation_report is not None:
        payload["validation_report"] = validation_report
    return payload


def serialize_staged_record(record: ImportStagedRecord) -> dict[str, Any]:
    def _load(text: str | None) -> Any:
        if not text:
            return []
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            return []

    return {
        "id": str(record.id),
        "row_number": record.row_number,
        "source_row_ref": record.source_row_ref,
        "raw_data": _load(record.raw_data),
        "mapped_data": _load(record.mapped_data),
        "validation_status": record.validation_status,
        "validation_errors": _load(record.validation_errors),
        "validation_warnings": _load(record.validation_warnings),
        "duplicate_status": record.duplicate_status,
        "duplicate_of": _load(record.duplicate_of),
        "reconciliation_status": record.reconciliation_status,
        "reconciliation_details": _load(record.reconciliation_details),
        "trust_level": record.trust_level,
        "promoted": record.promoted,
        "promoted_record_id": str(record.promoted_record_id) if record.promoted_record_id else None,
    }


def record_lineage(db: Session, entity_type: str, record_id: str) -> dict[str, Any] | None:
    """Trace a production record back to its source import (§26).

    Returns provenance info plus the originating job summary, or None when the
    entity/id is unknown.
    """
    import uuid as _u

    try:
        rid = _u.UUID(record_id)
    except ValueError:
        return None
    model = _model_for(entity_type)
    obj = db.query(model).filter(model.id == rid).first()
    if obj is None:
        return None
    lineage = {
        "entity_type": entity_type,
        "record_id": str(obj.id),
        "dataset_provenance": getattr(obj, "dataset_provenance", "live"),
        "source_import_job_id": str(obj.source_import_job_id) if getattr(obj, "source_import_job_id", None) else None,
        "source_file": getattr(obj, "source_file", None),
        "source_row_ref": getattr(obj, "source_row_ref", None),
        "created_at": obj.created_at.isoformat() if obj.created_at else None,
    }
    if lineage["source_import_job_id"]:
        job = db.query(ImportJob).filter(ImportJob.id == obj.source_import_job_id).first()
        if job is not None:
            lineage["import_job"] = serialize_job(job)
    return lineage
